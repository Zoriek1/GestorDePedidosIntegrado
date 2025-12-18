# -*- coding: utf-8 -*-
"""
Script para exportar vendas (pedidos) para planilha Excel
Agrupa por data de criação e calcula totais
"""
import sys
import os

# Adiciona o diretório backend ao path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from datetime import datetime, date
from decimal import Decimal
import re

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    os.system('pip install openpyxl')
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

from app import create_app, db
from app.models.pedido import Pedido


def parse_valor(valor_str):
    """Converte string de valor para float"""
    if not valor_str:
        return 0.0
    # Remove R$, espaços e troca vírgula por ponto
    valor_limpo = re.sub(r'[R$\s]', '', str(valor_str)).replace(',', '.')
    try:
        return float(valor_limpo)
    except:
        return 0.0


def exportar_vendas(data_inicio=None, data_fim=None, output_file=None):
    """
    Exporta pedidos para Excel
    
    Args:
        data_inicio: Data inicial (datetime ou string YYYY-MM-DD)
        data_fim: Data final (datetime ou string YYYY-MM-DD)
        output_file: Nome do arquivo de saída (padrão: vendas_YYYYMMDD.xlsx)
    """
    app = create_app()
    
    with app.app_context():
        # Monta a query
        query = Pedido.query.order_by(Pedido.created_at.desc())
        
        if data_inicio:
            if isinstance(data_inicio, str):
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(Pedido.created_at >= data_inicio)
        
        if data_fim:
            if isinstance(data_fim, str):
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            # Inclui o dia inteiro
            data_fim = datetime.combine(data_fim.date(), datetime.max.time())
            query = query.filter(Pedido.created_at <= data_fim)
        
        pedidos = query.all()
        
        if not pedidos:
            print("Nenhum pedido encontrado no período.")
            return None
        
        # Cria o workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        money_format = 'R$ #,##0.00'
        
        # Cabeçalhos
        headers = [
            'ID', 'Data Criação', 'Cliente', 'Destinatário', 'Produto',
            'Valor', 'Data Entrega', 'Horário', 'Status', 'Pagamento',
            'Status Pgto', 'Fonte', 'Cidade', 'Bairro'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        
        # Dados
        total_geral = 0.0
        for row, pedido in enumerate(pedidos, 2):
            valor = parse_valor(pedido.valor)
            total_geral += valor
            
            dados = [
                pedido.id,
                pedido.created_at.strftime('%d/%m/%Y %H:%M') if pedido.created_at else '',
                pedido.cliente,
                pedido.destinatario,
                pedido.produto,
                valor,
                pedido.dia_entrega.strftime('%d/%m/%Y') if pedido.dia_entrega else '',
                pedido.horario,
                pedido.status,
                pedido.pagamento or '',
                pedido.status_pagamento or '',
                pedido.fonte_pedido_rel.nome if pedido.fonte_pedido_rel else (pedido.fonte_pedido or ''),
                pedido.cidade or '',
                pedido.bairro or ''
            ]
            
            for col, valor_cell in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor_cell)
                cell.border = border
                if col == 6:  # Coluna Valor
                    cell.number_format = money_format
        
        # Linha de total
        row_total = len(pedidos) + 2
        ws.cell(row=row_total, column=1, value="TOTAL")
        ws.cell(row=row_total, column=1).font = Font(bold=True)
        
        ws.cell(row=row_total, column=5, value=f"{len(pedidos)} pedidos")
        ws.cell(row=row_total, column=5).font = Font(bold=True)
        
        cell_total = ws.cell(row=row_total, column=6, value=total_geral)
        cell_total.font = Font(bold=True)
        cell_total.number_format = money_format
        
        # Ajusta largura das colunas
        col_widths = [6, 18, 25, 25, 30, 12, 12, 10, 15, 15, 12, 15, 15, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Nome do arquivo
        if not output_file:
            output_file = f"vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Salva na pasta backups
        output_path = os.path.join(os.path.dirname(__file__), '..', 'backups', output_file)
        wb.save(output_path)
        
        print(f"\n✓ Exportado: {output_path}")
        print(f"  Total de pedidos: {len(pedidos)}")
        print(f"  Valor total: R$ {total_geral:,.2f}")
        
        return output_path


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Exportar vendas para Excel')
    parser.add_argument('--inicio', '-i', help='Data inicial (YYYY-MM-DD)')
    parser.add_argument('--fim', '-f', help='Data final (YYYY-MM-DD)')
    parser.add_argument('--output', '-o', help='Nome do arquivo de saída')
    
    args = parser.parse_args()
    
    print("=" * 50)
    print("EXPORTAÇÃO DE VENDAS PARA EXCEL")
    print("=" * 50)
    
    if args.inicio:
        print(f"Data início: {args.inicio}")
    if args.fim:
        print(f"Data fim: {args.fim}")
    
    exportar_vendas(args.inicio, args.fim, args.output)
